from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill


REPORT_COLUMNS = [
    'Casa',
    'Serie medidor',
    'Serial nuevo',
    'Ubicación',
    'Lectura actual',
    'Lectura anterior',
    'Consumo kWh',
    'Fecha inicial',
    'Fecha final',
    'Días',
    'Valor kWh',
    'Consumo en pesos',
    'Impuesto alumbrado 15%',
    'Total factura',
]


def build_report_excel(rows: list[dict], totals: dict) -> bytes:
    export_rows = rows.copy()
    export_rows.append(
        {
            'Casa': 'TOTAL',
            'Serie medidor': '',
            'Serial nuevo': '',
            'Ubicación': '',
            'Lectura actual': '',
            'Lectura anterior': '',
            'Consumo kWh': totals.get('Consumo kWh', 0),
            'Fecha inicial': '',
            'Fecha final': '',
            'Días': '',
            'Valor kWh': '',
            'Consumo en pesos': totals.get('Consumo en pesos', 0),
            'Impuesto alumbrado 15%': totals.get('Impuesto alumbrado 15%', 0),
            'Total factura': totals.get('Total factura', 0),
        }
    )

    df = pd.DataFrame(export_rows, columns=REPORT_COLUMNS)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Reporte')
        sheet = writer.sheets['Reporte']

        header_fill = PatternFill(start_color='4A6B3E', end_color='4A6B3E', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        total_fill = PatternFill(start_color='E8F0D8', end_color='E8F0D8', fill_type='solid')

        for col_idx, col in enumerate(REPORT_COLUMNS, start=1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            sheet.column_dimensions[cell.column_letter].width = max(16, len(col) + 2)

        total_row_idx = len(export_rows) + 1
        for col_idx in range(1, len(REPORT_COLUMNS) + 1):
            cell = sheet.cell(row=total_row_idx, column=col_idx)
            cell.fill = total_fill
            cell.font = Font(bold=True)

    output.seek(0)
    return output.read()
